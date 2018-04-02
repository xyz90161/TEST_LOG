# -*- coding: utf-8 -*-
'''
Created on 2017年9月26日

@author: equalhsiao
'''
import json
import os
import sqlite3
import sys
sys.path.append('D:\\test\\IL_TEST_LOG')
from flask import Flask
from flask.globals import session, request
from flask.templating import render_template
from openpyxl.reader.excel import load_workbook

from log.getLog import Tools





app = Flask(__name__)
app.secret_key = "super secret key"


def shutdown_server():
    func = request.environ.get('werkzeug.server.shutdown')
    if func is None:
        raise RuntimeError('Not running with the Werkzeug Server')
    func()
@app.route("/getBuildNumber",methods=['POST'])
def buildNumber():
    try:
        buildNumber = os.environ['BUILD_NUMBER']
    except:
        buildNumber = "0"
    return str(buildNumber)
@app.route("/shutdown",methods=['POST'])
def shutdown():
    shutdown_server()
    return 'Server shutting down...'
@app.route("/getDetail",methods=['POST'])
def getDetail():
#     data = request.get_json()
#     tool = Tools()
#     jsonSheetName = data['sheetName']
#     print jsonSheetName
#     index = data['index']
#     print index
#     nextIndex = data['nextIndex']
#     print nextIndex
#     wb3 = load_workbook(r"./test.xlsx")
#     for sheetName in wb3.get_sheet_names():
#         if jsonSheetName == sheetName:
#             worksheet = wb3.get_sheet_by_name(sheetName)
#             parameterList = tool.getCellParameter(int(index),int(nextIndex),worksheet)
#             break
#     print "parameterList:",parameterList
#     return parameterList[0]
#     
    pass
    
@app.route("/save", methods=['POST'])
def save():
    conn = sqlite3.connect('./../test.db')
    try:
        conn.execute('''CREATE TABLE TESTCASE
                       (ID INTEGER PRIMARY KEY AUTOINCREMENT,
                       SHEET          TEXT    NOT NULL,
                       ITEM           TEXT    NOT NULL,
                      STEP          TEXT     NOT NULL,
                      BUILDNUMBER    TEXT     NOT Null,
                      DEPEND        TEXT);''')
        conn.commit()
        
    except:
        print "TESTCASE table exits"
#         conn.execute('''DELETE FROM TESTCASE;''')
            
    conn.commit()
    data = request.get_json()
#     print data
    data_keys = data.keys()
    try:
        buildNumber = os.environ['BUILD_NUMBER']
    except:
        buildNumber = "0"
    for step,key in enumerate(data_keys):
        sheet_key = data[key].keys()
        for step2,k in enumerate(sheet_key):
            if len(data[key][k])>0:
                for item in data[key][k]:
                    if item is not None:
                        conn.execute("INSERT INTO TESTCASE (SHEET,ITEM,STEP,BUILDNUMBER,DEPEND) VALUES (?,?,?,?,?)",(k,item["name"],item['step'],buildNumber,item['depend']));
                        conn.commit()
    
#     buildUrl = os.environ['BUILD_URL']
    print " build number :",buildNumber
#     "build url :",buildUrl , 
    return str(buildNumber)
@app.route('/')
def index(name=None):
    print "index load excel path:",os.path.abspath('.')
    wb2 = load_workbook(r"./test.xlsx")
    tool = Tools()
    sheetList = []
    for index,sheetName in enumerate(wb2.get_sheet_names()):
        worksheet = wb2.get_sheet_by_name(sheetName)
        item = None
        item = tool.getItemColumn(worksheet)
        #取得項目的參數
        for cell in item:
            parameterList,cell_log =  tool.getCellParameter(cell['index'],cell['nextIndex'],worksheet)
            cell['parameterList'] = parameterList
        for i in item:
            i['sheetName'] = sheetName
        sheetList.append({"seq":index,"sheetName":sheetName,"item":item})
    session['sheetList'] = sheetList
    return render_template('processXLSX.html', name=name)
   
    
if __name__ == "__main__":
    app.debug = True
    app.run()