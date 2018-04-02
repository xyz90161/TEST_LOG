# -*- coding: utf-8 -*-
'''
Created on 2017年9月26日

@author: equalhsiao
'''

import logging
import os
import shutil
import sqlite3
import sys
sys.path.append('D:\\test\\IL_TEST_LOG')
import time
import unittest

from  openpyxl  import  load_workbook,Workbook 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from suds.bindings import binding
from suds.xsd.doctor import ImportDoctor, Import
import xmlrunner




#測試使用的方法
class Tools:
    #等待初始化完成
    def waitElement(self,driver,elemenType,elementAttr,waittime=10):
        try:
            element = WebDriverWait(driver, waittime).until(
                EC.presence_of_element_located((elemenType, elementAttr))
            )
            return element
        except Exception as e:
            print str(e)
            raise
    #建立webservice使用的tag
    def getParameter(self,client,name,value):
        insParameter = client.factory.create("ns0:Parameter")
        insParameter._name = name
        insParameter._value = value
        return insParameter
    #取得要執行的欄位A(名稱)、欄位B(項目)、欄位C(描述)
    def getItemColumn(self,worksheet):
        columnList = []
        tmpIndex = 0
        for row in range(2,worksheet.max_row): 
            for column in "ABC": 
                cell_name = "{}{}".format(column, row)
                if worksheet[cell_name].value is not None:
                    columnList.append({"value":worksheet[cell_name].value.strip(),"name":cell_name,"index":int(row),"prevIndex":tmpIndex})
                    tmpIndex = row
        for columnA in columnList:
            if "A" in columnA['name'] :
                for columnB in columnList:
                    if "B" in columnB['name']:
                        if columnB['index'] == columnA['index']:
                            columnA['requestType'] = columnB['value']
                            columnList.remove(columnB)
        for columnA in columnList:
            if "A" in columnA['name'] :
                for columnC in columnList:
                    if "C" in columnC['name']:
                        if columnC['index'] == columnA['index']:
                            columnA['description'] = columnC['value']
#                             .encode('utf8')
                            columnList.remove(columnC)
        tmpIndex = 0
        for i in range(0,len(columnList)):
            cell_item = columnList[i]
            try:
                if columnList[i+1]['index'] is not None:
                    cell_item['nextIndex']=columnList[i+1]['index']
            except IndexError:
                cell_item['nextIndex'] = worksheet.max_row
        return columnList
    
    #取得欄位D(參數名稱)與欄未E(參數值)存入 parameterList     
    def getCellParameter(self,startIndex,endIndex,worksheet):
        parameterList = []
        cell_log = ""
        for index in range(startIndex,endIndex+1): 
            for column in "D": 
                cell_name = "{}{}".format(column, index)
                for valueColum in "E":
                    cell2_name= "{}{}".format(valueColum,index)
                    if worksheet[cell2_name].value is not None and worksheet[cell_name].value is not None and worksheet[cell_name] != "":
                        name = str(worksheet[cell_name].value).strip()
                        value = str(worksheet[cell2_name].value).strip()
                        cell_log += "name:"+name+"value:"+value
                        parameterList.append({"name":name,"value":value})
            
        return parameterList,cell_log
    def getCheckLogCell(self,worksheet):
        for row in range(2,worksheet.max_row): 
            for column in "FGHI": 
                cell_name = "{}{}".format(column, row)
                print(cell_name)
    def getColumnCell(self,column,worksheet):
        columnList = []
        for index in range(1,worksheet.max_row):
            for column in column:
                cell_name = "{}{}".format(column,index)
                columnList.append(worksheet[cell_name])
        return columnList
    
        
    
    def callILWebService(self,parameterList,cell):
        from suds.client import Client
        wsdl = 'http://1.2.3.4:5/aaaa/bbbb?wsdl'
        headers = {'Content-Type': 'application/soap+xml'}
    #     proxy = dict(http='http://balala:lalala@fetfw.fareastone.com.tw:8080',
    #              https='http://balala:lalala@fetfw.fareastone.com.tw:8080')
        imp = Import('http://1.2.3.4:5678/aaaa/bbbb?xsd=1')
        d = ImportDoctor(imp)
        binding.envns = ('SOAP-ENV', 'http://www.w3.org/2003/05/soap-envelope')
        client = Client(wsdl,proxy = None,headers = headers,doctor = d)
        topRequest = None
        reqType =  str(cell['requestType'])
        print_log = ""
        print_log += "requestType:"+reqType
        if reqType == "CREATE":
            topRequest = client.factory.create("ns0:CreateRequest")
        elif reqType == "MODIFY":
            topRequest = client.factory.create("ns0:ModifyRequest")
        elif reqType == "DELETE":
            topRequest = client.factory.create("ns0:DeleteRequest")
        elif reqType == "DISPLAY":
            topRequest = client.factory.create("ns0:DisplayRequest")
        try:
            topRequest.RequestHeader.NeType = os.environ["NETYPE"]
            topRequest.RequestHeader.OrderNo = os.environ["ORDERNO"]
            topRequest.RequestHeader.ReqUser = os.environ["REQUSER"]
        except:
            topRequest.RequestHeader.NeType = "OMNI_BST"
            topRequest.RequestHeader.OrderNo = "1234"
            topRequest.RequestHeader.ReqUser = "COH"
        del topRequest.RequestHeader.MaxReqTime
        for parameter in parameterList:
                topRequest.RequestParameters.Parameter.append(self.getParameter(client,parameter['name'],parameter['value']))
        if reqType == "CREATE":
            response = client.service.create(topRequest.RequestHeader,topRequest.RequestParameters)
        elif reqType == "MODIFY":
            response = client.service.modify(topRequest.RequestHeader,topRequest.RequestParameters)
        elif reqType == "DELETE":
            response = client.service.delete(topRequest.RequestHeader,topRequest.RequestParameters)
        elif reqType == "DISPLAY":
            response = client.service.display(topRequest.RequestHeader,topRequest.RequestParameters)
#         response = client.service.create(topRequest.RequestHeader,topRequest.RequestParameters)
#         print response
        print_log += "SOAP Request >"+client.last_sent().__str__()+"\n"
        print_log += "SOAP Response >"+client.last_received().__str__()+"\n"
        if response is not None:
            return response.ResponseHeader.RequestId,response.ResponseHeader.Status,print_log

requestList = []
requestId = None
responseStatus = None
checkStatusList = []
log = ""
#測試IL的log是否正確
class test_checkLogUnit(unittest.TestCase):
    def test_chcekLog(self):
        logging.basicConfig( stream=sys.stderr )
        logging.getLogger( "test_checkLogUnit.test_chcekLog"+str(requestId) ).setLevel( logging.INFO )
        print log
        for checkObj in checkStatusList:
            checkString = checkObj['check']
            checkStatus = checkObj['status']
            self.assertFalse(checkStatus,"RequestId:"+str(requestId)+" checkstring:"+checkString)
#測試IL回覆的狀態是否為9
class test_ResponseUnitTest(unittest.TestCase):
    def test_ResponseStatus(self):
        logging.basicConfig( stream=sys.stderr )
        logging.getLogger( "test_ResponseUnitTest.test_ResponseStatus"+str(requestId)).setLevel( logging.INFO )
        print "requestId:",str(requestId)," responseStatus:",str(responseStatus)
        print log
        self.assertEqual("9", responseStatus, "requestId:"+str(requestId)+" responseStatus:"+str(responseStatus))
    def tearDown(self):
        pass
    
if __name__ == '__main__':
        #logging.basicConfig(level=logging.INFO)
        print "test"
#         logging.getLogger('suds.client').setLevel(logging.DEBUG)
#         logging.getLogger('suds.transport').setLevel(logging.DEBUG)
#         logging.getLogger('suds.xsd.schema').setLevel(logging.DEBUG)
#         logging.getLogger('suds.wsdl').setLevel(logging.DEBUG)
    #初始化建立目錄
        if os.path.exists("./unittest-reports"):
            shutil.rmtree("./unittest-reports")
        tool = Tools()
        conn = sqlite3.connect('../test.db')
        cur = conn.cursor()
        try:
            buildNumber = os.environ['BUILD_NUMBER']
            print "build number:",buildNumber
        except:
            buildNumber = "0"
        cur.execute("select * from TESTCASE where BUILDNUMBER = ? order by STEP ",(buildNumber,));
        testCaseAll = cur.fetchall();
        #讀取excel
        try:
            print "getLog load excel path:",os.path.abspath('.')
            wb2 = load_workbook(r"./test.xlsx")
        except:
            print "[load ./../test/test.xlsx for test]"
            wb2 = load_workbook(r"../test/test.xlsx")
        requestList = []
        wb = Workbook()
        ws = wb.active
        for sheetName in wb2.get_sheet_names():
            ws.title = sheetName
            worksheet = wb2.get_sheet_by_name(sheetName)
            #取得要執行的項目
            ActionList = tool.getItemColumn(worksheet)
            newActionList = []
            for testCase in testCaseAll:
                for Action in ActionList:
                    if testCase[2] == Action['value']:
                        Action['seq'] = int(testCase[3])
                        newActionList.append(Action)
            newActionList.sort(key = lambda x:x.get('seq'))
            for step,cell in enumerate(newActionList):
                ws.cell(row=step, column=1, value=cell['value'])
                currentStatus = False
                dependStatus = False
                log = ""
                for testCase in testCaseAll:
                    if sheetName == testCase[1]:
                        if  cell['value'] in testCase[2]:
                            currentStatus = True
                if not currentStatus:
                    continue
#                 dependCur = conn.cursor()
#                 dependCur.execute("select DEPEND from TESTCASE where BUILDNUMBER = ? and ITEM = ?",(buildNumber,cell['value']))
#                 depend =  dependCur.fetchone()
#                 if depend is not None:
#                     print "depend:",depend
                
#                 print "cell",cell
                            
                #取得項目的參數
                parameterList,cell_log =  tool.getCellParameter(cell['index'],cell['nextIndex'],worksheet)
#                 log += cell_log
#                 log = ""+cell.__str__()
                log += " Action item:"+str(cell['value'])+"\n"
#                 log += parameterList.__str__()+"\n"
#                 log += "--------------------------------Action item:"+str(cell['value'])+"  [parameter list  end]-------------------------------------------\n"
                checkList = []
                #取得檢查的條件
                for check in list(worksheet.rows)[int(cell['index']-1)]:
                    if (check.column in "ABC"):
                        continue
                    if (check.value is not None):
                        checkList.append(check.value.encode("utf8",'ignore').strip())
                #call IL 的webservice
                requestId,responseStatus,print_log = tool.callILWebService(parameterList,cell)
                log += print_log
                #值型單元測式並取產生xml的報表
                suite = unittest.TestLoader().loadTestsFromTestCase(test_ResponseUnitTest)
                xmlrunner.XMLTestRunner(output="unittest-reports",outsuffix=str(requestId)).run(suite)
                requestList.append({"requestId":str(requestId),"checkList":checkList})
        wb.save('result.xlsx')
        exit(0)
        #取得IL的帳號密碼
        try:
            userName =  os.environ['IL_USERNAME']
            password = os.environ['IL_PASSWORD']
            buildType = os.environ['BUILD_TYPE']
        except:
            #預設的帳號密碼
            userName = "equalhsiao"
            password = "123456"
            buildType = "webdriver"
        if "webdriver" in buildType:
            driver = webdriver.Firefox(executable_path="D:\\geckodriver.exe")
        elif "phantomjs" in buildType:
            driver = webdriver.PhantomJS(executable_path=r"D:\tools\phantomjs-2.1.1-windows\bin\phantomjs.exe")
        #登入
        driver.get("http://1.2.3.4:5678/aaa/index.jsp?sessionExpired=1")
        tool.waitElement(driver,By.NAME, "username")
        tool.waitElement(driver,By.NAME, "password")
        tool.waitElement(driver,By.NAME, "btnauthenticate")
        driver.find_element_by_name("username").send_keys(userName)
        driver.find_element_by_name("password").send_keys(password)
        driver.find_element_by_tag_name("body").send_keys(Keys.ESCAPE)
        driver.find_element_by_name("btnauthenticate").click()
        frame_main = tool.waitElement(driver, By.XPATH, "//frame[@name='main']")
        driver.switch_to_frame(frame_main)
        tool.waitElement(driver, By.CLASS_NAME, "header")
        header = driver.find_element_by_class_name("header").text
        print "debug username:"+str(userName)+" password:"+str(password[0]+"*"*len(password)+password[len(password)-1:])+"\n"
        #檢查log是否包含測試的字串
        for requestObject in requestList:
            log = ""    
            requestId = requestObject['requestId']
            checkList = requestObject['checkList']
            if "Welcome to Comptel InstantLink" in header:
                driver.get("http://1.2.3.4:5678/aaaa/bbbb/ccccc?request=2&requestId="+requestId+"&logType=mml&specificTaskId=&openOrCloseAllTaskLogs=OPEN_ALL")
            else:
                log += "login error\n"
                driver.quit()
            preList = driver.find_elements_by_tag_name("pre")
#             errorStatus = True
            checkStatusList = []
            for check in checkList:
                checkStatusList.append({"check":check,"status":True})
            #若有包含則便更狀態為False
            #False表示正確,True表是錯誤
            for pre in preList:
                log += pre.text.replace("<","").encode("utf8")
                print pre.text.replace("<","").encode("utf8")
                for checkObj in checkStatusList:
                    check = checkObj['check']
                    if check in pre.text:
                        checkObj['status'] = False
            log += "\n"+"checkList:"+checkList.__str__()
            #執行單元測試
            new_suite = unittest.TestLoader().loadTestsFromTestCase(test_checkLogUnit)
            xmlrunner.XMLTestRunner(output='unittest-reports',outsuffix=str(requestId)).run(new_suite)
            time.sleep(1)
        driver.close()
        driver.quit()
